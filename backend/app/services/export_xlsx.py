"""Build Excel workbook for OKR export. Objectives sheet + Key Results sheet."""
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from io import BytesIO


def _serialize_cell(val):
    if val is None:
        return ''
    if hasattr(val, 'isoformat'):
        return val.isoformat()
    if isinstance(val, (list, tuple)):
        return ', '.join(str(x) for x in val)
    return str(val)


def build_okr_workbook(db, objectives, out_stream):
    """Write an xlsx workbook to out_stream. objectives is list of raw MongoDB objective docs."""
    wb = Workbook()
    ws_obj = wb.active
    ws_obj.title = 'Objectives'
    obj_headers = [
        '_id', 'title', 'description', 'ownerId', 'level', 'timeline', 'fiscalYear', 'quarter',
        'parentObjectiveId', 'division', 'status', 'departmentId', 'createdAt', 'updatedAt',
    ]
    ws_obj.append(obj_headers)
    for row_idx, doc in enumerate(objectives, start=2):
        row = [_serialize_cell(doc.get(h)) for h in obj_headers]
        ws_obj.append(row)

    # Key Results sheet
    ws_kr = wb.create_sheet('Key Results', 1)
    kr_headers = [
        '_id', 'objectiveId', 'title', 'target', 'currentValue', 'unit', 'score', 'targetScore',
        'ownerId', 'createdAt', 'lastUpdatedAt',
    ]
    ws_kr.append(kr_headers)
    for doc in objectives:
        oid = doc.get('_id')
        if not oid:
            continue
        for kr in db.key_results.find({'objectiveId': oid}):
            row = [_serialize_cell(kr.get(h)) for h in kr_headers]
            ws_kr.append(row)

    # Style header row
    for ws in (ws_obj, ws_kr):
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
    wb.save(out_stream)
